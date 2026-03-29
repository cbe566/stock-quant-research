#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
結構型產品報價單 Excel 生成器
根據已成單報價資料，生成格式化的 Excel 報價單
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from copy import copy

def create_product_sheet(wb, sheet_name, product_data):
    """建立單一產品的報價單工作表"""
    ws = wb.create_sheet(title=sheet_name)

    # 定義樣式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 字體
    bold_font = Font(name='Microsoft JhengHei', bold=True, size=11)
    normal_font = Font(name='Microsoft JhengHei', size=11)
    red_font = Font(name='Microsoft JhengHei', bold=True, size=11, color='FF0000')
    white_bold_font = Font(name='Microsoft JhengHei', bold=True, size=11, color='FFFFFF')
    green_font = Font(name='Microsoft JhengHei', bold=True, size=11, color='006400')

    # 填充色
    light_green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    dark_green_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    light_gray_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    header_green_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')

    # 對齊
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # 設定欄寬
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    # === 上半部：產品參數 ===
    params = [
        ('Type', '類型', product_data['type']),
        ('Tenor', '天期', product_data['tenor']),
        ('Strike', '預計執行價', product_data['strike']),
        ('KO', '出場價', product_data['ko']),
        ('Coupon', '年化報酬率', product_data['coupon']),
        ('KO Start', '閉鎖', product_data['ko_start']),
        ('Memory KO', '記憶式', product_data['memory_ko']),
        ('Currency', '幣別', product_data['currency']),
    ]

    for i, (eng, chi, value) in enumerate(params, start=1):
        row = i

        # A 欄：英文名稱（淺綠底）
        cell_a = ws.cell(row=row, column=1, value=eng)
        cell_a.font = bold_font
        cell_a.fill = light_green_fill
        cell_a.border = thin_border
        cell_a.alignment = left_align

        # B 欄：中文名稱
        cell_b = ws.cell(row=row, column=2, value=chi)
        cell_b.font = normal_font
        cell_b.fill = white_fill
        cell_b.border = thin_border
        cell_b.alignment = center_align

        # C+D 欄合併：數值（紅色粗體）
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        cell_c = ws.cell(row=row, column=3, value=value)
        cell_c.font = red_font
        cell_c.fill = white_fill
        cell_c.border = thin_border
        cell_c.alignment = center_align
        # 合併區域的邊框
        for col in range(3, 6):
            ws.cell(row=row, column=col).border = thin_border

    # === 發行機構列 ===
    row_issuer = len(params) + 1
    ws.merge_cells(start_row=row_issuer, start_column=1, end_row=row_issuer, end_column=5)
    cell_issuer = ws.cell(row=row_issuer, column=1, value=f"發行機構　{product_data['issuer']}")
    cell_issuer.font = Font(name='Microsoft JhengHei', bold=True, size=11, color='006400')
    cell_issuer.fill = light_green_fill
    cell_issuer.border = thin_border
    cell_issuer.alignment = left_align
    for col in range(1, 6):
        ws.cell(row=row_issuer, column=col).border = thin_border

    # === 連結標的表頭 ===
    row_header = row_issuer + 1
    headers = ['連結標的', '', '參考進場價', '預計執行價', '提前出場價']
    sub_headers = ['', '', product_data['ref_date'], product_data['strike'], product_data['ko']]

    # 第一行表頭
    ws.merge_cells(start_row=row_header, start_column=1, end_row=row_header, end_column=2)
    cell_h1 = ws.cell(row=row_header, column=1, value='連結標的')
    cell_h1.font = white_bold_font
    cell_h1.fill = dark_green_fill
    cell_h1.border = thin_border
    cell_h1.alignment = center_align
    ws.cell(row=row_header, column=2).border = thin_border
    ws.cell(row=row_header, column=2).fill = dark_green_fill

    header_labels = ['參考進場價', '預計執行價', '提前出場價']
    header_subs = [product_data['ref_date'], product_data['strike'], product_data['ko']]
    for j, (label, sub) in enumerate(zip(header_labels, header_subs), start=3):
        cell = ws.cell(row=row_header, column=j, value=f"{label}\n{sub}")
        cell.font = white_bold_font
        cell.fill = dark_green_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 設定表頭行高
    ws.row_dimensions[row_header].height = 35

    # === 連結標的數據 ===
    for i, underlying in enumerate(product_data['underlyings']):
        row_data = row_header + 1 + i
        ticker, name, ref_price, strike_price, ko_price = underlying

        # 標的代號
        cell_ticker = ws.cell(row=row_data, column=1, value=ticker)
        cell_ticker.font = bold_font
        cell_ticker.fill = light_gray_fill if i % 2 == 0 else white_fill
        cell_ticker.border = thin_border
        cell_ticker.alignment = left_align

        # 中文名稱
        cell_name = ws.cell(row=row_data, column=2, value=name)
        cell_name.font = normal_font
        cell_name.fill = light_gray_fill if i % 2 == 0 else white_fill
        cell_name.border = thin_border
        cell_name.alignment = center_align

        # 參考進場價
        cell_ref = ws.cell(row=row_data, column=3, value=ref_price)
        cell_ref.font = normal_font
        cell_ref.fill = light_gray_fill if i % 2 == 0 else white_fill
        cell_ref.border = thin_border
        cell_ref.alignment = center_align
        cell_ref.number_format = '#,##0.00'

        # 預計執行價
        cell_strike = ws.cell(row=row_data, column=4, value=strike_price)
        cell_strike.font = normal_font
        cell_strike.fill = light_gray_fill if i % 2 == 0 else white_fill
        cell_strike.border = thin_border
        cell_strike.alignment = center_align
        cell_strike.number_format = '#,##0.00'

        # 提前出場價
        cell_ko = ws.cell(row=row_data, column=5, value=ko_price)
        cell_ko.font = normal_font
        cell_ko.fill = light_gray_fill if i % 2 == 0 else white_fill
        cell_ko.border = thin_border
        cell_ko.alignment = center_align
        cell_ko.number_format = '#,##0.00'


def main():
    wb = openpyxl.Workbook()
    # 移除預設工作表
    wb.remove(wb.active)

    # === 產品 1：2026SN1183 STEPDOWNFCN ===
    product1 = {
        'type': 'STEPDOWN FCN(固定配息)',
        'tenor': '',
        'strike': '75.00%',
        'ko': '100%',
        'coupon': '12.02%',
        'ko_start': '1',
        'memory_ko': 'YES',
        'currency': 'JPY',
        'issuer': 'DSU',
        'ref_date': '2026/3/27',
        'underlyings': [
            ('AMD', '超微半導體', 201.99, 151.49, 201.99),
            ('NVDA', '輝達', 167.52, 125.64, 167.52),
            ('TSM', '台積電', 326.74, 245.06, 326.74),
            ('GOOGL', '谷歌', 274.34, 205.76, 274.34),
        ]
    }
    create_product_sheet(wb, 'SN1183_STEPDOWNFCN', product1)

    # === 產品 2：2026SN1184 FCN ===
    product2 = {
        'type': 'FCN(固定配息)',
        'tenor': '',
        'strike': '',
        'ko': '100%',
        'coupon': '67.36%',
        'ko_start': '1',
        'memory_ko': 'YES',
        'currency': 'USD',
        'issuer': 'DSU',
        'ref_date': '2026/3/27',
        'underlyings': [
            ('TSLA', '特斯拉', 361.83, None, 361.83),
            ('NVDA', '輝達', 167.52, None, 167.52),
            ('AMD', '超微半導體', 201.99, None, 201.99),
            ('INTC', '英特爾', 43.13, None, 43.13),
        ]
    }
    create_product_sheet(wb, 'SN1184_FCN', product2)

    # === 產品 3：2026SN1185 FCN ===
    product3 = {
        'type': 'FCN(固定配息)',
        'tenor': '',
        'strike': '',
        'ko': '100%',
        'coupon': '71.52%',
        'ko_start': '1',
        'memory_ko': 'YES',
        'currency': 'USD',
        'issuer': 'DSU',
        'ref_date': '2026/3/27',
        'underlyings': [
            ('TSM', '台積電', 326.74, None, 326.74),
            ('NVDA', '輝達', 167.52, None, 167.52),
            ('AVGO', '博通', 300.68, None, 300.68),
            ('AAL', '美國航空', 10.30, None, 10.30),
        ]
    }
    create_product_sheet(wb, 'SN1185_FCN', product3)

    # 儲存檔案
    output_path = '/Users/jamie/Desktop/Claude-股票量化數據研究網/結構型產品報價單_已成單.xlsx'
    wb.save(output_path)
    print(f'✅ Excel 報價單已生成：{output_path}')
    print(f'📊 共 {len(wb.sheetnames)} 個工作表：{", ".join(wb.sheetnames)}')


if __name__ == '__main__':
    main()
